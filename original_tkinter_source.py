# -*- coding: utf-8 -*-







"""







Капилляриметрия — интерактивный GUI v4.1 (FULL)







Полный функционал:







- Авто/ручной ρ (по минерализации или вручную), мгновенный пересчёт всего







- Таблица ступеней (P, Вкл., Масса, R) редактируется в приложении







- P–S (drag по X) с фикс. узлом S=1 на P=0 (из исходников), сетка, подсказки







- R–P (drag по Y) с инверсией P: слева 10 → справа 0, точка R0 ромбом, сетка, подсказки







- Цвета: измерено — зелёный; интерп. — оранжевый; вручную (drag) — синий







- Плато масс: если 10 атм не измерена, m(10)=m(7)







- Kво по последней включённой ступени: Kво = (m_sat - m_last) / (m_sat - m_dry)







- Экспорт в Excel (openpyxl) или CSV







"""















import math







import tkinter as tk







from tkinter import ttk, messagebox, filedialog







from typing import List, Tuple, Optional















# === Константы ===







PRESSURES = [0.25, 0.5, 1, 3, 5, 7, 10]  # атм















# Matplotlib








# === ОБОЗНАЧЕНИЯ СОПРОТИВЛЕНИЙ ===
# R₀  — сопротивление образца при 100% насыщении (Ом·м)
# Rt  — сопротивление образца при текущем насыщении (Ом·м)
# Rw  — удельное сопротивление модельной пластовой воды (Ом·м)
# R   — измеренное сопротивление на ступени (Ом·м)
# Rcorr — скорректированное сопротивление (с приведением по Rw и R₀)
#
# Все сопротивления должны быть заданы в одной системе — Ом·м.
#
# === ОСНОВНЫЕ ФОРМУЛЫ ===
# φ = (a * Rw / R₀)^(1/m)                — пористость (Pп)
# S_w = ((a * Rw) / (φ^m * Rt))^(1/n)    — водонасыщенность (Pн)
# I = Rt / R₀                            — индекс насыщения (безразмерный)

import matplotlib







matplotlib.use("TkAgg")







from matplotlib.figure import Figure







from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg















# Excel (опционально)







HAS_XLSX = True







try:







    from openpyxl import Workbook







except Exception:







    HAS_XLSX = False















# Цвета







CLR_MEAS = "#2e7d32"   # измерено







CLR_INTR = "#ef6c00"   # интерп.







CLR_MAN  = "#1565c0"   # вручную (drag)







CLR_FIRST= "#000000"   # S=1,P=0 (на P–S)







CLR_R0   = "#1976d2"   # R0 ромб







BG_AXES  = "#f2f2f2"















# ---------- Утилиты ----------







def r4(x: float) -> float:







    try: return round(float(x), 4)







    except Exception: return float('nan')















# ===== OСТ 39-204-86 helpers (added) =====







g_OST = 9.80665  # m/s^2















def _ost_apply_salt_correction_41(S_res, C_percent, gamma_b, gamma_p, gamma_o):







    """







    ОСТ 39-204-86, формула (4.1): поправка остаточной водонасыщенности на соли.







    S_res: доля (0..1); C_percent: % масс. (в 100 г воды); плотности в г/см3.







    """







    try:







        C = float(C_percent); S = float(S_res)







        eps = 1e-12







        denom1 = (100.0 - C) if abs(100.0 - C) > eps else eps







        numer = gamma_b * (1.0 + C / denom1)







        denom = gamma_p * (1.0 + (S * C * gamma_b) / (gamma_o * denom1 if abs(gamma_o) > eps else eps))







        return S * (numer / denom)







    except Exception:







        return float('nan')















def _ost_pc_to_H_51(Pc_lab_Pa, sigma_lab, sigma_pl, delta_gamma_pl, cos_theta_ratio=1.0):







    """







    Практическая форма 5.1: Pc_pl = Pc_lab * (σ_pl/σ_lab) * cosθ_ratio; H = Pc_pl / (Δρ_pl * g).







    Все входы в СИ. Возвращает H (м).







    """







    try:







        Pc_pl = Pc_lab_Pa * (sigma_pl / (sigma_lab if sigma_lab else 1e-12)) * float(cos_theta_ratio)







        return Pc_pl / (delta_gamma_pl * g_OST if delta_gamma_pl else 1e-12)







    except Exception:







        return float('nan')







# =========================================















def build_cubic_spline(x, y):







    """Natural cubic spline по массивам x (возраст.) и y."""







    n = len(x)







    if n < 2: return None







    a = y[:]







    h = [x[i+1] - x[i] for i in range(n-1)]







    alpha = [0.0]*(n-1)







    for i in range(1, n-1):







        alpha[i] = (3.0/h[i])*(a[i+1]-a[i]) - (3.0/h[i-1])*(a[i]-a[i-1])







    l = [0.0]*n; mu = [0.0]*n; z = [0.0]*n







    l[0] = 1.0; mu[0] = 0.0; z[0] = 0.0







    for i in range(1, n-1):







        l[i] = 2.0*(x[i+1]-x[i-1]) - h[i-1]*mu[i-1]







        mu[i] = h[i]/l[i]







        z[i] = (alpha[i] - h[i-1]*z[i-1]) / l[i]







    l[n-1] = 1.0; z[n-1] = 0.0







    c = [0.0]*n; b = [0.0]*(n-1); d = [0.0]*(n-1)







    for j in range(n-2, -1, -1):







        c[j] = z[j] - mu[j]*c[j+1]







        b[j] = (a[j+1]-a[j])/h[j] - h[j]*(c[j+1] + 2.0*c[j])/3.0







        d[j] = (c[j+1]-c[j])/(3.0*h[j])







    return {"x": x[:-1], "a": a[:-1], "b": b, "c": c[:-1], "d": d}















def eval_spline(sp, xq: float) -> float:







    if sp is None: return float('nan')







    xs, a, b, c, d = sp["x"], sp["a"], sp["b"], sp["c"], sp["d"]







    i = 0







    if xq <= xs[0]: i = 0







    elif xq >= xs[-1]: i = len(xs)-1







    else:







        for k in range(len(xs)-1):







            if xs[k] <= xq <= xs[k+1]:







                i = k; break







    dx = xq - xs[i]







    return a[i] + b[i]*dx + c[i]*dx*dx + d[i]*dx*dx*dx















def lin_interp(x: float, pts: List[Tuple[float, float]]) -> float:







    """Линейная интерполяция по (x, y); x монотонен."""







    if not pts: return float('nan')







    if len(pts) == 1: return pts[0][1]







    pts = sorted(pts, key=lambda t: t[0])







    if x <= pts[0][0]:







        (x1, y1), (x2, y2) = pts[0], pts[1]







    elif x >= pts[-1][0]:







        (x1, y1), (x2, y2) = pts[-2], pts[-1]







    else:







        x1=y1=x2=y2=None







        for i in range(len(pts)-1):







            if pts[i][0] <= x <= pts[i+1][0]:







                x1, y1 = pts[i]; x2, y2 = pts[i+1]; break







    if x1 is None or x2 is None or x2 == x1: return float('nan')







    return y1 + (x - x1)*((y2 - y1)/(x2 - x1))















# ---------- Приложение ----------







class App(tk.Tk):







    def __init__(self):







        super().__init__()







        self.title("Капилляриметрия — интерактив v4.1 (FULL)")







        self.geometry("1520x920")







        self.minsize(1280, 820)















        style = ttk.Style(self); style.theme_use("clam")







        style.configure("Treeview", rowheight=28, borderwidth=1, relief="solid")







        style.configure("Treeview.Heading", borderwidth=1, relief="solid")







        style.map("Treeview", background=[("selected", "#cfe8ff")])















        paned = ttk.Panedwindow(self, orient=tk.HORIZONTAL); paned.pack(fill="both", expand=True)







        left = ttk.Frame(paned); right = ttk.Frame(paned)







        paned.add(left, weight=3); paned.add(right, weight=4)















        # ---- Исходные данные ----







        inputs = ttk.LabelFrame(left, text="Исходные данные"); inputs.pack(fill="x", padx=8, pady=6)







        self.vars = {







            "D": tk.StringVar(), "L": tk.StringVar(),







            "rho": tk.StringVar(value="1.0000"),







            "m_dry": tk.StringVar(), "m_sat": tk.StringVar(),







            "R0": tk.StringVar(),







            "salinity": tk.StringVar(value=""),







            "auto_rho": tk.BooleanVar(value=True)







        }







        self._add_entry(inputs, "D (мм):", "D", 0, 0)







        self._add_entry(inputs, "L (мм):", "L", 0, 2)















        ttk.Checkbutton(inputs, variable=self.vars["auto_rho"],







                        text="Авто ρ по минерализации (г/л NaCl)",







                        command=self._on_toggle_auto_rho).grid(row=0, column=4, padx=6, pady=4, sticky="w")







        self._add_entry(inputs, "Минерализация (г/л):", "salinity", 1, 0)















        ttk.Label(inputs, text="ρ (г/см³):").grid(row=1, column=2, padx=6, pady=4, sticky="e")







        self.entry_rho = ttk.Entry(inputs, width=12, textvariable=self.vars["rho"])







        self.entry_rho.grid(row=1, column=3, padx=6, pady=4, sticky="w")







        # --- Поле для ручного ввода удельного сопротивления воды (ρв, Ом·м) ---







        ttk.Label(inputs, text='ρв — сопротивление воды (Ом·м):').grid(row=15, column=0, padx=5, pady=2, sticky='e')







        self.vars['R_water'] = tk.StringVar(value='1.0')







        self.entry_Rw = ttk.Entry(inputs, width=12, textvariable=self.vars['R_water'])







        self.entry_Rw.grid(row=15, column=1, padx=5, pady=2, sticky='w')







        self.entry_Rw.bind('<FocusOut>', lambda *_: self._recalc_if_ready())







        self.entry_Rw.bind('<Return>', lambda *_: self._recalc_if_ready())















                # Пересчёт при изменении плотности вручную (фокус/Enter)







        self.entry_rho.bind('<FocusOut>', lambda *_: self._recalc_if_ready())







        self.entry_rho.bind('<Return>', lambda *_: self._recalc_if_ready())















        self._add_entry(inputs, "m сух (г):", "m_dry", 2, 0)







        self._add_entry(inputs, "m нас (г):", "m_sat", 2, 2)







        self._add_entry(inputs, "R₀ — при 100% насыщении (Ом·м):", "R0", 2, 4)







        # Пересчёт при изменении R0 (фокус/Enter)







        try:







            self.entries['R0'].bind('<FocusOut>', lambda *_: self._recalc_if_ready())







            self.entries['R0'].bind('<Return>', lambda *_: self._recalc_if_ready())







        except Exception:







            pass















        # live bindings







        for key in ("D","L","m_dry","m_sat","R0"):







            self._bind_live_recalc(self.vars[key])







        self._bind_live_recalc(self.vars["salinity"], is_salinity=True)







        self._bind_live_recalc(self.vars["rho"], is_rho=True)







        self._bind_live_recalc(self.vars["R0"], is_rho=True)







        self._bind_live_recalc(self.vars["R_water"], is_rho=True)







        self._on_toggle_auto_rho(initial=True)















        # ---- Таблица ступеней ----







        steps = ttk.LabelFrame(left, text="Ступени давления и измерения (редактируемо)")







        steps.pack(fill="both", expand=False, padx=8, pady=6)







        self.tree_inputs = ttk.Treeview(steps, columns=("P","en","m","R"), show="headings", height=8)







        for col, text, w in (("P","P (атм)",80), ("en","Вкл.",60), ("m","Масса (г)",140), ("R","R (Ом)",140)):







            self.tree_inputs.heading(col, text=text); self.tree_inputs.column(col, width=w, anchor="center")







        self.tree_inputs.pack(fill="x", padx=6, pady=6)







        self.tree_inputs.tag_configure("edited", background="#e1ecff")















        self.step_data = []







        for p in PRESSURES:







            self.step_data.append({"P":p,"en":True,"m":"","R":""})







            self.tree_inputs.insert("", "end", values=(p, "✓", "", ""))















        # --- Archie parameters block (auto-inserted) ---
        try:
            import numpy as np  # ensure numpy available in this scope
        except Exception:
            pass

        # Archie constants (can be moved to GUI inputs)
        a = 1.0
        m = 2.0

        # Collect points to fit n from ln(I) = -n ln(Sw), where I = rho/rho0
        Sw_list = []
        rho_list = []  # using 'R' as resistivity (Ohm*m) per user workflow
        for row in getattr(self, "step_data", []):
            try:
                Sw = float(row.get("Sw", 0))
                rho = float(row.get("R", 0))
                if Sw > 0 and rho > 0:
                    Sw_list.append(Sw)
                    rho_list.append(rho)
            except Exception:
                pass

        n_value = None
        if len(Sw_list) >= 2:
            rho0 = rho_list[0]  # reference as first valid point
            lnS = np.log(np.asarray(Sw_list, dtype=float))
            lnI = np.log(np.asarray(rho_list, dtype=float) / float(rho0))
            coeffs = np.polyfit(lnS, lnI, 1)  # ln I = m*ln S + b, here m ~ -n
            n_value = -float(coeffs[0])

        # Fill back per-step: I, Pp (from Kпор), Pn, and n
        for row in getattr(self, "step_data", []):
            try:
                Sw = float(row.get("Sw", 0))
                rho = float(row.get("R", 0))
                phi = float(row.get("Kпор", 0)) if "Kпор" in row else float(row.get("Kpor", 0))
                if Sw > 0 and rho > 0:
                    rho0 = rho_list[0] if rho_list else None
                    row["I"] = (rho / rho0) if rho0 else None
                    row["Pp"] = (a * (phi ** (-m))) if (phi and phi > 0) else None
                    row["Pn"] = (Sw ** (-n_value)) if (n_value is not None and Sw > 0) else None
                    row["n"]  = n_value
            except Exception:
                pass
        # --- End of Archie block ---
        self.tree_inputs.bind("<Double-1>", self._start_edit_cell)







        self.tree_inputs.bind("<Button-1>", self._toggle_en_if_clicked)















        btns = ttk.Frame(steps); btns.pack(fill="x", padx=6, pady=4)







        ttk.Button(btns, text="Сбросить графики", command=self.reset_curves).pack(side="right", padx=4)







        ttk.Button(btns, text="Экспорт", command=self.export_results).pack(side="right", padx=4)















        # ---- Результаты ----







        out = ttk.LabelFrame(left, text="Результаты"); out.pack(fill="both", expand=True, padx=8, pady=6)







        self.summary = tk.Text(out, height=8, wrap="word"); self.summary.pack(fill="x", padx=6, pady=6)







        self.tree_res = ttk.Treeview(out,







            columns=("P","m","mSrc","R","rSrc","Vv","S","S_pct","lam","I","log_Sw","log_I","n"),







            show="headings", height=12)







        headers = [("P","P (атм)",70), ("m","Масса (г)",100), ("mSrc","Источник m",110),







                   ("R","R (Ом)",100), ("rSrc","Источник R",110), ("Vv","Vв.выт (см³)",110),







                   ("S","S (доля)",90), ("S_pct","S (%)",80), ("lam","λ=R0/R",90), ("I","I (индекс)",90), ("log_Sw","log(Sw)",90), ("log_I","log(I)",90), ("n","n (Арчи)",90)]







        for c,t,w in headers:







            self.tree_res.heading(c, text=t); self.tree_res.column(c, width=w, anchor="center")







        self.tree_res.pack(fill="both", expand=True, padx=6, pady=6)















        # ---- Графики ----







        charts = ttk.LabelFrame(right, text="Графики (перетаскивание точек)")







        charts.pack(fill="both", expand=True, padx=8, pady=6)















        self.fig1 = Figure(figsize=(6.8,4.2), dpi=100); self.ax1 = self.fig1.add_subplot(111, facecolor=BG_AXES)







        self.canvas1 = FigureCanvasTkAgg(self.fig1, master=charts); self.canvas1.get_tk_widget().pack(fill="both", expand=True, padx=6, pady=(6,0))







        self._build_legend(charts, [("Измерено", CLR_MEAS, "o"), ("Интерп.", CLR_INTR, "o"),







                                    ("Ручн. (drag)", CLR_MAN, "o"), ("Насыщ. (S=1,P=0)", CLR_FIRST, "o")])















        self.fig2 = Figure(figsize=(6.8,4.2), dpi=100); self.ax2 = self.fig2.add_subplot(111, facecolor=BG_AXES)







        self.canvas2 = FigureCanvasTkAgg(self.fig2, master=charts); self.canvas2.get_tk_widget().pack(fill="both", expand=True, padx=6, pady=(6,0))







        self._build_legend(charts, [("Измерено", CLR_MEAS, "o"), ("Интерп.", CLR_INTR, "o"),







                                    ("Ручн. (drag)", CLR_MAN, "o"), ("R0 (P=0)", CLR_R0, "D")])















        # состояния







        self.ps_vline = self.ps_hline = self.ps_tooltip = None







        self.rp_vline = self.rp_hline = self.rp_tooltip = None







        self.dragging = None







        self.last_rows = None; self.meta = None







        self.baseS = self.baseR = self.baseR0 = None







        self.curS = self.curR = self.curR0 = None







        self.m_type = self.r_type = None







        self.dragged_flag = False















        # события графиков







        self.canvas1.mpl_connect('motion_notify_event', self._on_motion_ps_crosshair)







        self.canvas1.mpl_connect('axes_leave_event', self._on_leave_ps)







        self.canvas1.mpl_connect('button_press_event', self._on_press_ps)







        self.canvas1.mpl_connect('button_release_event', self._on_release_ps)







        self.canvas1.mpl_connect('motion_notify_event', self._on_drag_ps)















        self.canvas2.mpl_connect('motion_notify_event', self._on_motion_rp_crosshair)







        self.canvas2.mpl_connect('axes_leave_event', self._on_leave_rp)







        self.canvas2.mpl_connect('button_press_event', self._on_press_rp)







        self.canvas2.mpl_connect('button_release_event', self._on_release_rp)







        self.canvas2.mpl_connect('motion_notify_event', self._on_drag_rp)















        self._draw_empty()















    # ---------- UI helpers ----------







    def _add_entry(self, parent, label, key, row, col):







        ttk.Label(parent, text=label).grid(row=row, column=col, padx=6, pady=4, sticky="e")







        e = ttk.Entry(parent, width=12, textvariable=self.vars[key])







        e.grid(row=row, column=col+1, padx=6, pady=4, sticky="w")







        e.bind("<FocusOut>", lambda *_: self._recalc_if_ready())















    def _bind_live_recalc(self, tk_var: tk.Variable, is_salinity=False, is_rho=False):







        def on_change(*_):







            # Реакция на изменение минерализации или плотности







            if is_salinity:







                if self.vars['auto_rho'].get():







                    # Автоматический пересчёт ρ по минерализации







                    self._update_rho_from_salinity()







                else:







                    # Ручной режим — ρ фиксировано, но пересчёт по минерализации всё равно выполняем







                    try:







                        self._recalc_if_ready()







                    except Exception as e:







                        print('Ошибка пересчёта по минерализации:', e)







            elif is_rho:







                try:







                    self._recalc_if_ready()







                except Exception as e:







                    print('Ошибка пересчёта по плотности:', e)







            else:







                self._recalc_if_ready()







        tk_var.trace_add('write', lambda *_: on_change())















    def _on_toggle_auto_rho(self, initial=False):







        auto = self.vars["auto_rho"].get()







        self.entry_rho.configure(state=("readonly" if auto else "normal"))







        if auto:







            self._update_rho_from_salinity()







        if not initial:







            self._recalc_if_ready()















    def _update_rho_from_salinity(self):







        C = self._try_float(self.vars["salinity"].get())







        if C is None: C = 0.0







        # ОСТ 39-204-86: C в г/л; диапазон 0..300 г/л (по ТЗ пользователя)







        C = max(0.0, min(300.0, C))







        # ρ = 0.9982 + 0.0008 * C  (г/см³)







        rho = 0.9982 + 0.0008 * C







        rho = max(0.6, min(1.5, rho))







        self.vars["rho"].set(f"{rho:.4f}")







        # Для поправки ОСТ 4.1 нужна соль в % на 100 г воды: C_percent = 0.1 * C







        try:







            self.meta["ost_C_percent"] = 0.1 * C







        except Exception:







            pass















    def _build_legend(self, parent, items):







        box = ttk.Frame(parent); box.pack(fill="x", padx=6, pady=(2,10))







        for text, color, marker in items:







            f = ttk.Frame(box); f.pack(side="left", padx=8)







            sw = tk.Canvas(f, width=16, height=16, bg="white", highlightthickness=0); sw.pack(side="left")







            if marker == "o":







                sw.create_oval(3,3,13,13, fill=color, outline=color)







            elif marker == "D":







                sw.create_polygon(8,2,14,8,8,14,2,8, fill=color, outline=color)







            ttk.Label(f, text=text).pack(side="left", padx=4)















    def _draw_empty(self):







        self.ax1.clear()







        self.ax1.set_title("P vs S (drag по X)")







        self.ax1.set_xlabel("S (доля)"); self.ax1.set_ylabel("P (атм)")







        self.ax1.set_xlim(0,1); self.ax1.grid(True, linestyle="--", linewidth=0.5, alpha=0.6)







        self.canvas1.draw()















        self.ax2.clear()







        self.ax2.set_title("R vs P (drag по Y)")







        self.ax2.set_xlabel("P (атм)"); self.ax2.set_ylabel("R (Ом)")







        self.ax2.set_xlim(10,0)  # инверсия по X: слева 10 → справа 0







        self.ax2.grid(True, linestyle="--", linewidth=0.5, alpha=0.6)







        self.canvas2.draw()















    # ---------- Таблица ступеней ----------







    def _treeview_identify(self, event):







        region = self.tree_inputs.identify("region", event.x, event.y)







        row_id = self.tree_inputs.identify_row(event.y)







        col_id = self.tree_inputs.identify_column(event.x)







        return region, row_id, col_id















    def _toggle_en_if_clicked(self, event):







        region, row_id, col_id = self._treeview_identify(event)







        if region == "cell" and col_id == "#2" and row_id:







            idx = self.tree_inputs.index(row_id)







            self.step_data[idx]["en"] = not self.step_data[idx]["en"]







            self._refresh_inputs_table()







            self._recalc_if_ready()















    def _start_edit_cell(self, event):







        region, row_id, col_id = self._treeview_identify(event)







        if region != "cell" or not row_id or col_id not in ("#3", "#4"):







            return







        x, y, w, h = self.tree_inputs.bbox(row_id, col_id)







        idx = self.tree_inputs.index(row_id)







        col_key = "m" if col_id == "#3" else "R"







        old = self.step_data[idx][col_key]















        editor = ttk.Entry(self.tree_inputs)







        editor.place(x=x, y=y, width=w, height=h)







        editor.insert(0, old); editor.focus_set()















        def commit(*_):







            val = editor.get().strip()







            self.step_data[idx][col_key] = val







            editor.destroy()







            item = self.tree_inputs.get_children()[idx]







            self.tree_inputs.item(item, tags=("edited",))







            self._refresh_inputs_table()







            self._recalc_if_ready()















        def cancel(*_): editor.destroy()















        editor.bind("<Return>", commit)







        editor.bind("<FocusOut>", commit)







        editor.bind("<Escape>", cancel)















    def _refresh_inputs_table(self):







        for i in self.tree_inputs.get_children():







            self.tree_inputs.delete(i)







        for d in self.step_data:







            self.tree_inputs.insert("", "end",







                                    values=(d["P"], "✓" if d["en"] else "", d["m"], d["R"]),







                                    tags=("edited",) if d.get("edited", False) else ())















    # ---------- Пересчёт ----------







    def _try_float(self, s: str) -> Optional[float]:







        s = (s or "").strip()







        if not s: return None







        try: return float(s.replace(",", "."))







        except Exception: return None















    def _recalc_if_ready(self):







        try:







            _ = [float(self.vars[k].get().replace(",", ".")) for k in ("D","L","rho","m_dry","m_sat")]







        except Exception:







            return







        self.calculate(auto=True)















    def calculate(self, auto: bool=False):







        try:







            D  = float(self.vars["D"].get().replace(",", "."))







            L  = float(self.vars["L"].get().replace(",", "."))







            rho= float(self.vars["rho"].get().replace(",", "."))







            m_dry = float(self.vars["m_dry"].get().replace(",", "."))







            m_sat = float(self.vars["m_sat"].get().replace(",", "."))







        except Exception:







            if not auto:







                messagebox.showerror("Ошибка", "Заполни D, L, ρ, m_сух и m_нас корректными числами.")







            return















        R0v = self._try_float(self.vars["R0"].get())







        masses = []; resistances = []; enabled = []







        for d in self.step_data:







            masses.append(self._try_float(d["m"]))







            resistances.append(self._try_float(d["R"]))







            enabled.append(bool(d["en"]))















        Vobr = math.pi * (D/2.0)**2 * L * 1e-3







        m_water_total = m_sat - m_dry







        Vp = m_water_total / rho if rho != 0 else float('nan')















        # Массы: сплайн по log10(P)







        mass_pts = [(math.log10(p), m) for p, m, en in zip(PRESSURES, masses, enabled) if en and m is not None and p>0]







        mass_pts.sort(key=lambda t: t[0])







        sp = build_cubic_spline([t[0] for t in mass_pts], [t[1] for t in mass_pts]) if len(mass_pts) >= 2 else None















        final_m = []; m_src = []; self.m_type = []







        for p, m, en in zip(PRESSURES, masses, enabled):







            if en and m is not None:







                final_m.append(m); m_src.append("измерено"); self.m_type.append("meas")







            elif p>0 and sp is not None:







                final_m.append(eval_spline(sp, math.log10(p))); m_src.append("интерп."); self.m_type.append("intr")







            else:







                final_m.append(float('nan')); m_src.append("—"); self.m_type.append("intr")















        # Плато 10 = 7, если 10 не измерена







        i10 = PRESSURES.index(10); i7 = PRESSURES.index(7)







        if m_src[i10] != "измерено" and not math.isnan(final_m[i7]):







            final_m[i10] = final_m[i7]; m_src[i10] = "интерп.(плато)"; self.m_type[i10] = "intr"















        # R: линейно по P (+ R0 при 0)







        rpts = []







        if R0v is not None: rpts.append((0.0, R0v))







        for p, r, en in zip(PRESSURES, resistances, enabled):







            if en and r is not None: rpts.append((float(p), r))







        rpts.sort(key=lambda t: t[0])















        final_R = []; r_src = []; self.r_type = []







        for p, r, en in zip(PRESSURES, resistances, enabled):







            if en and r is not None:







                final_R.append(r); r_src.append("измерено"); self.r_type.append("meas")







            else:







                Rp = lin_interp(float(p), rpts) if rpts else float('nan')







                final_R.append(Rp); r_src.append("интерп." if not math.isnan(Rp) else "—"); self.r_type.append("intr")



















                        # Собираем строки результата

        rows = []

        Sarr = []

        for i, p in enumerate(PRESSURES):

            m = final_m[i]

            R = final_R[i]

            # Объём вытесненной воды и насыщенность по массе

            Vv = (m_sat - m)/rho if (isinstance(m, float) and not math.isnan(m) and rho not in (None, 0.0)) else float('nan')

            S  = ((Vp - Vv)/Vp) if (not math.isnan(Vv) and isinstance(Vp, float) and Vp > 0) else float('nan')
            # lam по измеренному сопротивлению образца (100% / текущая ступень)

            lam = ((R0v / R) if (R0v is not None and R is not None and not math.isnan(R) and R > 0) else float('nan'))

            rows.append({

                "P": p,

                "m": m,

                "mSrc": m_src[i],

                "R": R,

                "rSrc": r_src[i],

                "Vv": Vv,

                "S": S,

                "S_pct": (S*100.0 if not math.isnan(S) else float('nan')),

                "lam": lam

            })

            Sarr.append(S)

# Сохраняем состояние







        self.meta = {"D":D,"L":L,"rho":rho,"m_dry":m_dry,"m_sat":m_sat,"R0":R0v,







                     "Vobr":Vobr,"Vp":Vp,"enabled":enabled}







        self.last_rows = rows







        if self.baseS is None:







            self.baseS = Sarr[:]; self.baseR = final_R[:]; self.baseR0 = R0v







        self.curS = Sarr[:]; self.curR = final_R[:]; self.curR0 = R0v







        self.dragged_flag = False















        self._render_all(rows, dragged=False)















    # ---------- Отрисовка ----------







    def _render_all(self, rows, dragged: bool):







        self._render_results(rows)







        self._render_summary_with_kvo(dragged)







        self._draw_charts()















    def _render_results(self, rows):

        # --- добавляем расчёт Арчи ---
        import numpy as np, math
        if rows and isinstance(rows, list) and len(rows) > 1:
            Sw = np.array([float(r.get("S", r.get("S_pct", 0))/100 if "S_pct" in r else r.get("S", np.nan)) for r in rows])
            R = np.array([float(r.get("R", np.nan)) for r in rows])
            try:
                idx_R0 = np.nanargmin(np.abs(Sw - 1))
                R0 = R[idx_R0]
            except Exception:
                R0 = np.nan
            I = R / R0 if R0 and R0 != 0 else np.full_like(R, np.nan)
            with np.errstate(divide='ignore', invalid='ignore'):
                log_Sw = np.log10(Sw)
                log_I = np.log10(I)
            mask = np.isfinite(log_Sw) & np.isfinite(log_I)
            n_val = -np.polyfit(log_Sw[mask], log_I[mask], 1)[0] if mask.sum() >= 2 else np.nan
            for i, r in enumerate(rows):
                r["I"] = float(I[i]) if np.isfinite(I[i]) else None
                r["log_Sw"] = float(log_Sw[i]) if np.isfinite(log_Sw[i]) else None
                r["log_I"] = float(log_I[i]) if np.isfinite(log_I[i]) else None
                r["n"] = float(n_val) if np.isfinite(n_val) else None
        # --- конец вставки Арчи ---








        for i in self.tree_res.get_children(): self.tree_res.delete(i)







        for r in rows:







            def fmt(v):







                if v is None: return ""







                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return ""







                return f"{v:.6f}" if isinstance(v, float) else str(v)







            self.tree_res.insert("", "end",







                values=(r["P"], fmt(r["m"]), r["mSrc"], fmt(r["R"]), r["rSrc"],







                        fmt(r["Vv"]), fmt(r["S"]), fmt(r["S_pct"]), fmt(r["lam"]), fmt(r.get("I")), fmt(r.get("log_Sw")), fmt(r.get("log_I")), fmt(r.get("n"))))















    def _render_summary_with_kvo(self, dragged: bool):







        # --- Пористость по ОСТ 39-204-86: последняя ступень давления ---







        try:







            if getattr(self, "last_rows", None):







                last = self.last_rows[-1]







                Vv_max = float(last.get("Vv", 0.0))







                Vobr = self.meta.get("Vobr")







                if Vobr and Vobr > 0:







                    self.meta["Kpor"] = Vv_max / Vobr







                    # --- поправка: эффективная пористость с учетом остаточной водонасыщенности ---







                    try:







                        S_last = float(last.get("S", float('nan')))







                        if not (isinstance(S_last, float) and not math.isnan(S_last)):







                            S_last = float('nan')







                    except Exception:







                        S_last = float('nan')







                    # Попробуем применить ОСТ 4.1, если есть минерализация







                    S_res_corr = None







                    try:







                        C_percent = self.meta.get("ost_C_percent")







                        gamma_b = self.meta.get("rho")  # плотность воды/раствора, г/см3







                        gamma_p = self.meta.get("rho")  # нет отдельного ввода, используем rho как приближение







                        gamma_o = self.meta.get("rho_o", self.meta.get("rho"))







                        if (C_percent is not None) and (gamma_b is not None) and (gamma_p is not None) and (gamma_o is not None) and isinstance(S_last, float) and not math.isnan(S_last):







                            S_res_corr = _ost_apply_salt_correction_41(S_last, C_percent, gamma_b, gamma_p, gamma_o)







                    except Exception:







                        S_res_corr = None







                    # Если коррекцию посчитать не удалось — берём как есть







                    if not (isinstance(S_res_corr, float) and not math.isnan(S_res_corr)):







                        S_res_corr = S_last







                    # Сохраняем в meta для отображения







                    try:







                        if isinstance(S_res_corr, float) and not math.isnan(S_res_corr):







                            self.meta["S_res_OST41"] = max(0.0, min(1.0, S_res_corr))







                    except Exception:







                        pass







                    Kpor_eff = (Vv_max / Vobr) * (1.0 - (self.meta.get("S_res_OST41") if isinstance(self.meta.get("S_res_OST41"), float) else (S_last if isinstance(S_last,float) else 0.0)))







        except Exception:







            pass







        # -------------------------------------------------------------















        # OST per-row additions: Pc_Pa and H_m







        try:







            unit = self.meta.get("pc_units", "atm")







            def _to_Pa(P):







                if unit == "atm": return P * 101325.0







                if unit == "bar": return P * 1e5







                if unit == "kPa": return P * 1e3







                if unit == "MPa": return P * 1e6







                return P







            for _r in (self.last_rows or []):







                P = _r.get("P")







                if P is None: continue







                Pc_Pa = _to_Pa(float(P))







                _r["Pc_Pa"] = Pc_Pa







                _r["H_m"] = _ost_pc_to_H_51(







                    Pc_lab_Pa=Pc_Pa,







                    sigma_lab=self.meta.get("ost_sigma_lab", 0.072),







                    sigma_pl=self.meta.get("ost_sigma_pl", 0.030),







                    delta_gamma_pl=self.meta.get("ost_delta_gamma_pl", 200.0),







                )







        except Exception:







            pass















        m = self.meta; Vobr = m["Vobr"]; Vp = m["Vp"]







        rho, m_dry, m_sat = m["rho"], m["m_dry"], m["m_sat"]







        # Kво по последней включённой ступени







        k_last = None







        for idx in range(len(PRESSURES)-1, -1, -1):







            if m["enabled"][idx]: k_last = idx; break







        kvo_val = float('nan'); kvo_pct = float('nan')







        if k_last is not None and self.last_rows:







            m_last = self.last_rows[k_last]["m"]







            if m_last is not None and not math.isnan(m_last) and (m_sat - m_dry) != 0:







                kvo_val = max(0.0, min(1.0, (m_sat - m_last)/(m_sat - m_dry)))







                kvo_pct = kvo_val * 100.0







                # Остаточная водонасыщенность (Sов) по последней ступени







                sov_val = max(0.0, min(1.0, (m_last - m_dry)/(m_sat - m_dry)))







                sov_pct = sov_val * 100.0















        self.summary.delete("1.0", "end")







        self.summary.insert("end", f"V_обр = {r4(Vobr)} см³\nV_p = {r4(Vp)} см³\nm_вода(100%) = {r4(m_sat - m_dry)} г\n")







        self.summary.insert("end", f"ρ жидкости = {r4(rho)} г/см³\nR0(100%) = {('—' if self.curR0 is None else r4(self.curR0))} Ом\n")







        # Выбор Sов: сначала корректированная по ОСТ-41, если доступна







        sov_mass = (float("nan") if (m_sat - m_dry) == 0 else max(0.0, min(1.0, (m_last - m_dry)/(m_sat - m_dry)))) if (k_last is not None and self.last_rows and m_last is not None and not math.isnan(m_last)) else float("nan")







        sov_corr = self.meta.get("S_res_OST41", float("nan"))







        if isinstance(sov_corr, float) and not math.isnan(sov_corr):







            sov_val = sov_corr







        else:







            sov_val = sov_mass







        kvo_val = (float("nan") if math.isnan(sov_val) else max(0.0, min(1.0, 1.0 - sov_val)))







        kvo_pct = (float("nan") if math.isnan(kvo_val) else kvo_val*100.0)







        # Вывод







        self.summary.insert("end", f"Доля вытесненной воды = {('—' if math.isnan(kvo_val) else str(r4(kvo_val)))} ({'—' if math.isnan(kvo_pct) else str(r4(kvo_pct)) + '%'})\n")







        self.summary.insert("end", f"Sов (остаточная вода) = {('—' if math.isnan(sov_val) else str(r4(sov_val)))} ({'—' if math.isnan(sov_val) else str(r4(sov_val*100.0)) + '%'})\n")







        # Для прозрачности дополнительно показываем Sов по массам







        self.summary.insert("end", f"Sов по массам = {('—' if math.isnan(sov_mass) else str(r4(sov_mass)))} ({'—' if math.isnan(sov_mass) else str(r4(sov_mass*100.0)) + '%'})\n")







        self.summary.insert("end", f"Drag-редактирование: {'Да' if dragged or self.dragged_flag else 'Нет'}\n")







        self.kvo_val = kvo_val; self.kvo_pct = kvo_pct







        self.sov_val = sov_val; self.sov_pct = (float("nan") if math.isnan(sov_val) else sov_val*100.0)







        self.kvo_val = kvo_val; self.kvo_pct = kvo_pct







        self.sov_val = sov_val; self.sov_pct = sov_pct















    # ---------- Графики ----------







        try:







            self.summary.insert("end", f"\nKпор. = {self._fmt(self.meta.get('Kpor'))} (доля)\n")







            self.summary.insert("end", f"Kпор. эфф. = {self._fmt(self.meta.get('Kpor_eff'))} (доля)\n")







            self.summary.insert("end", f"S_в.ост(ОСТ 4.1) = {self._fmt(self.meta.get('S_res_OST41'))} (доля)\n")







        except Exception:







            pass















        try:







            self.summary.insert("end", f"Kпор. = {self._fmt(self.meta.get('Kpor'))} (доля)\n")







            self.summary.insert("end", f"Kпор. эфф. = {self._fmt(self.meta.get('Kpor_eff'))} (доля)\n")







            rho_used = self.vars.get("rho").get() if self.vars.get("rho") else ""







            mode = self.meta.get("rho_mode","auto_by_C" if self.vars.get("auto_rho").get() else "manual")







            self.summary.insert("end", f"ρ использовано = {rho_used} г/см³ [{mode}]\n")







        except Exception:







            pass















        # === Always append porosity and rho lines (visible in Results) ===







        try:







            kpor = self.meta.get('Kpor')







            kpor_eff = self.meta.get('Kpor_eff')







            kpor_pct = (kpor * 100.0) if (isinstance(kpor, (int,float)) and not math.isnan(kpor)) else None







            kpor_eff_pct = (kpor_eff * 100.0) if (isinstance(kpor_eff, (int,float)) and not math.isnan(kpor_eff)) else None







            self.summary.insert("end", f"Kпор. = {self._fmt(kpor)} (доля) / {self._fmt(kpor_pct)} %\n")







            self.summary.insert("end", f"Kпор. эфф. = {self._fmt(kpor_eff)} (доля) / {self._fmt(kpor_eff_pct)} %\n")







            rho_used = self.vars.get("rho").get() if hasattr(self, "vars") and self.vars.get("rho") else ""







            mode = "auto_by_C" if (hasattr(self, "vars") and self.vars.get("auto_rho") and self.vars["auto_rho"].get()) else self.meta.get("rho_mode","manual")







            self.summary.insert("end", f"ρ использовано = {rho_used} г/см³ [{mode}]\n")







        except Exception:







            pass







        # ===============================================================















        # --- Пористость по ОСТ 39-204-86 (по последней ступени давления) ---







        try:







            if getattr(self, "last_rows", None):







                last = self.last_rows[-1]







                Vv_max = float(last.get("Vv", 0.0))







                Vobr = self.meta.get("Vobr")







                if Vobr and Vobr > 0:







                    Kpor = Vv_max / Vobr







                    # --- поправка: эффективная пористость с учетом остаточной водонасыщенности ---







                    try:







                        S_last = float(last.get("S", float('nan')))







                        if not (isinstance(S_last, float) and not math.isnan(S_last)):







                            S_last = float('nan')







                    except Exception:







                        S_last = float('nan')







                    S_res_corr = None







                    try:







                        C_percent = self.meta.get("ost_C_percent")







                        gamma_b = self.meta.get("rho")







                        gamma_p = self.meta.get("rho")







                        gamma_o = self.meta.get("rho_o", self.meta.get("rho"))







                        if (C_percent is not None) and (gamma_b is not None) and (gamma_p is not None) and (gamma_o is not None) and isinstance(S_last, float) and not math.isnan(S_last):







                            S_res_corr = _ost_apply_salt_correction_41(S_last, C_percent, gamma_b, gamma_p, gamma_o)







                    except Exception:







                        S_res_corr = None







                    if not (isinstance(S_res_corr, float) and not math.isnan(S_res_corr)):







                        S_res_corr = S_last







                    try:







                        if isinstance(S_res_corr, float) and not math.isnan(S_res_corr):







                            self.meta["S_res_OST41"] = max(0.0, min(1.0, S_res_corr))







                    except Exception:







                        pass







                    Kpor_eff = Kpor * (1.0 - (self.meta.get("S_res_OST41") if isinstance(self.meta.get("S_res_OST41"), float) else (S_last if isinstance(S_last,float) else 0.0)))







                    self.meta["Kpor"] = Kpor







                    self.meta["Kpor_eff"] = Kpor_eff







                    self.summary.insert("end", f"Kпор. = {Kpor:.4f} (доля) / {Kpor*100:.2f} %\n")







                    self.summary.insert("end", f"Kпор. эфф. = {Kpor_eff:.4f} (доля) / {Kpor_eff*100:.2f} %\n")







        except Exception as e:







            print("Ошибка расчета Kпор:", e)







        # -------------------------------------------------------------















    def _color_for_m(self, idx):







        t = "intr" if self.m_type is None else self.m_type[idx]







        return CLR_INTR if t == "intr" else (CLR_MEAS if t == "meas" else CLR_MAN)















    def _color_for_r(self, idx):







        t = "intr" if self.r_type is None else self.r_type[idx]







        return CLR_INTR if t == "intr" else (CLR_MEAS if t == "meas" else CLR_MAN)















    def _draw_charts(self):







        # P–S (фикс узел S=1, P=0)







        self.ax1.clear(); self.ax1.set_facecolor(BG_AXES)







        xs = [1.0]; ys = [0.0]; colors = [CLR_FIRST]; sizes=[56]







        for i, p in enumerate(PRESSURES):







            s = self.curS[i]







            if s is None or (isinstance(s, float) and math.isnan(s)): continue







            xs.append(s); ys.append(p); colors.append(self._color_for_m(i)); sizes.append(56)







        self.ax1.plot(xs, ys, linestyle="-", color="#7f8c8d", linewidth=1.0)







        self.ps_scatter = self.ax1.scatter(xs, ys, s=sizes, c=colors, marker="o", zorder=3)







        self.ax1.set_xlim(0,1); self.ax1.set_xlabel("S (доля)"); self.ax1.set_ylabel("P (атм)")







        self.ax1.grid(True, linestyle="--", linewidth=0.5, alpha=0.6)







        self.canvas1.draw()















        # R–P (инверсия по X). Точка R0 на P=0 — ромб







        self.ax2.clear(); self.ax2.set_facecolor(BG_AXES)







        order = [10,7,5,3,1,0.5,0.25,0]







        ymap = {p: (self.curR[PRESSURES.index(p)] if p in PRESSURES else None) for p in PRESSURES}







        ymap[0.0] = self.curR0







        xs2, ys2, colors2, sizes2 = [], [], [], []







        for p in order:







            v = ymap.get(p, None)







            xs2.append(p)







            ys2.append(float('nan') if (v is None or (isinstance(v,float) and math.isnan(v))) else v)







            if p == 0:







                colors2.append(CLR_R0); sizes2.append(64)







            elif p in PRESSURES:







                colors2.append(self._color_for_r(PRESSURES.index(p))); sizes2.append(56)







            else:







                colors2.append("#7f8c8d"); sizes2.append(56)







        self.ax2.plot(xs2, ys2, linestyle="-", color="#7f8c8d", linewidth=1.0)







        self.rp_scatter = self.ax2.scatter(xs2, ys2, s=sizes2, c=colors2, marker="o", zorder=3)







        if ymap[0.0] is not None and not (isinstance(ymap[0.0],float) and math.isnan(ymap[0.0])):







            self.ax2.scatter([0],[ymap[0.0]], s=90, c=CLR_R0, marker="D", zorder=4)







        self.ax2.set_xlim(10,0); self.ax2.set_xlabel("P (атм)"); self.ax2.set_ylabel("R (Ом)")







        self.ax2.grid(True, linestyle="--", linewidth=0.5, alpha=0.6)







        self.canvas2.draw()















    # ---------- Кроссхэр/подсказки ----------







    def _ensure_ps_guides(self):







        if self.ps_vline is None: self.ps_vline = self.ax1.axvline(color="#888", lw=0.6, ls=":")







        if self.ps_hline is None: self.ps_hline = self.ax1.axhline(color="#888", lw=0.6, ls=":")







        if self.ps_tooltip is None:







            self.ps_tooltip = self.ax1.annotate("", xy=(0,0), xytext=(12,12),







                textcoords="offset points",







                bbox=dict(boxstyle="round,pad=0.3", fc="#ffffe6", ec="#888", lw=0.5),







                fontsize=9)







            self.ps_tooltip.set_visible(False)















    def _ensure_rp_guides(self):







        if self.rp_vline is None: self.rp_vline = self.ax2.axvline(color="#888", lw=0.6, ls=":")







        if self.rp_hline is None: self.rp_hline = self.ax2.axhline(color="#888", lw=0.6, ls=":")







        if self.rp_tooltip is None:







            self.rp_tooltip = self.ax2.annotate("", xy=(0,0), xytext=(12,12),







                textcoords="offset points",







                bbox=dict(boxstyle="round,pad=0.3", fc="#ffffe6", ec="#888", lw=0.5),







                fontsize=9)







            self.rp_tooltip.set_visible(False)















    def _nearest_ps_point(self, x, y):







        candidates = [(1.0, 0.0, -1)]







        for i,p in enumerate(PRESSURES):







            s = self.curS[i]







            if s is None or (isinstance(s,float) and math.isnan(s)): continue







            candidates.append((s,p,i))







        best = None; best_d2 = 1e9







        for sx,py,idx in candidates:







            dx = x - sx; dy = y - py; d2 = dx*dx + dy*dy







            if d2 < best_d2: best_d2 = d2; best = (sx,py,idx)







        return best if best_d2 < 0.5**2 else None















    def _nearest_rp_point(self, x, y):







        order = [10,7,5,3,1,0.5,0.25,0]







        pts = []







        for p in order:







            if p == 0: R = self.curR0







            else: R = self.curR[PRESSURES.index(p)]







            if R is None or (isinstance(R,float) and math.isnan(R)): continue







            pts.append((float(p), float(R), p))







        best = None; best_d2 = 1e9







        for px,py,tag in pts:







            dx = x - px; dy = y - py; d2 = dx*dx + dy*dy







            if d2 < best_d2: best_d2 = d2; best = (px,py,tag)







        return best if best_d2 < 0.5**2 else None















    # Crosshair + tooltip (PS)







    def _on_motion_ps_crosshair(self, event):







        if event.inaxes != self.ax1: return







        self._ensure_ps_guides()







        if event.xdata is None or event.ydata is None:







            self._on_leave_ps(None); return







        x, y = float(event.xdata), float(event.ydata)







        self.ps_vline.set_xdata([x, x]); self.ps_hline.set_ydata([y, y])















        pt = self._nearest_ps_point(x, y)







        if pt is not None:







            sx, py, idx = pt







            if idx == -1:







                m_txt = ""







                if self.meta: m_txt = f", m={r4(self.meta['m_sat'])} г"







                self.ps_tooltip.xy = (sx, py)







                self.ps_tooltip.set_text(f"P=0 атм\nS=1.0{m_txt}")







            else:







                s = self.curS[idx]; p = PRESSURES[idx]







                m_txt = ""







                if self.meta:







                    Vp = self.meta["Vp"]; rho = self.meta["rho"]; m_sat = self.meta["m_sat"]







                    if Vp and Vp>0:







                        Vv = Vp*(1-s); mval = m_sat - rho*Vv







                        m_txt = f", m={r4(mval)} г"







                self.ps_tooltip.xy = (s, p)







                self.ps_tooltip.set_text(f"P={r4(p)} атм\nS={r4(s)}{m_txt}")







        else:







            S = min(1.0, max(0.0, x)); P = y







            m_txt = ""







            if self.meta:







                Vp = self.meta["Vp"]; rho = self.meta["rho"]; m_sat = self.meta["m_sat"]







                if Vp and Vp>0:







                    Vv = Vp*(1*S); mval = m_sat - rho*Vv







                    m_txt = f", m={r4(mval)} г"







            self.ps_tooltip.xy = (x, y)







            self.ps_tooltip.set_text(f"P={r4(P)} атм\nS={r4(S)}{m_txt}")















        self.ps_tooltip.set_visible(True)







        self.canvas1.draw_idle()















    def _on_leave_ps(self, event):







        if self.ps_tooltip: self.ps_tooltip.set_visible(False)







        self.canvas1.draw_idle()















    # Crosshair + tooltip (RP)







    def _on_motion_rp_crosshair(self, event):







        if event.inaxes != self.ax2: return







        self._ensure_rp_guides()







        if event.xdata is None or event.ydata is None:







            self._on_leave_rp(None); return







        x, y = float(event.xdata), float(event.ydata)







        self.rp_vline.set_xdata([x, x]); self.rp_hline.set_ydata([y, y])















        pt = self._nearest_rp_point(x, y)







        if pt is not None:







            px, py, tag = pt







            lam_txt = ""







            if py>0 and self.curR0 is not None:







                lam_txt = f", λ={r4(self.curR0/py)}"







            self.rp_tooltip.xy = (px, py)







            self.rp_tooltip.set_text(f"P={r4(px)} атм\nR={r4(py)}{lam_txt}")







        else:







            P = x; R = max(0.0, y)







            lam_txt = ""







            if R>0 and self.curR0 is not None:







                lam_txt = f", λ={r4(self.curR0/R)}"







            self.rp_tooltip.xy = (x, y)







            self.rp_tooltip.set_text(f"P={r4(P)} атм\nR={r4(R)}{lam_txt}")















        self.rp_tooltip.set_visible(True)







        self.canvas2.draw_idle()















    def _on_leave_rp(self, event):







        if self.rp_tooltip: self.rp_tooltip.set_visible(False)







        self.canvas2.draw_idle()















    # ---------- Drag ----------







    def _nearest_ps_idx(self, y_value):







        best_k = None; best_d = 1e9







        for i,p in enumerate(PRESSURES):







            d = abs(p - y_value)







            if d < best_d: best_d, best_k = d, i







        return best_k















    def _on_press_ps(self, event):







        if self.curS is None or event.inaxes != self.ax1: return







        if event.xdata is None or event.ydata is None: return







        idx = self._nearest_ps_idx(event.ydata)







        if idx is not None:







            self.dragging = ("ps", idx)















    def _on_drag_ps(self, event):







        if not (self.dragging and self.dragging[0]=="ps"): return







        if event.inaxes != self.ax1 or event.xdata is None: return







        idx = self.dragging[1]







        newS = min(1.0, max(0.0, float(event.xdata)))







        self.curS[idx] = newS







        if self.m_type is None: self.m_type = ["intr"]*len(PRESSURES)







        self.m_type[idx] = "man"; self.dragged_flag = True







        # обновляем массу в таблице (учитывая текущие rho и Vp)







        if self.meta:







            Vp = self.meta["Vp"]; rho = self.meta["rho"]; m_sat = self.meta["m_sat"]







            if Vp and Vp>0:







                Vv = Vp*(1-newS); m = m_sat - rho*Vv







                self.step_data[idx]["m"] = f"{m:.6f}"







        self._recompute_from_curves(drag_source="S")







        self._draw_charts()















    def _on_release_ps(self, event):







        if self.dragging and self.dragging[0]=="ps":







            self.dragging = None







            self._mark_dragged()















    def _nearest_rp_idx(self, x_value):







        xs = [0] + PRESSURES[:]







        best_i = None; best_d = 1e9







        for i,x in enumerate(xs):







            d = abs(x - x_value)







            if d < best_d: best_d, best_i = d, i







        if best_d <= 0.6: return best_i







        return None















    def _on_press_rp(self, event):







        if self.curR is None or event.inaxes != self.ax2: return







        if event.xdata is None or event.ydata is None: return







        idx = self._nearest_rp_idx(event.xdata)







        if idx is not None:







            self.dragging = ("rp", idx)















    def _on_drag_rp(self, event):







        if not (self.dragging and self.dragging[0]=="rp"): return







        if event.inaxes != self.ax2 or event.ydata is None: return







        newR = max(0.0, float(event.ydata))







        idx = self.dragging[1]







        if idx == 0:







            self.curR0 = newR







            self.vars["R0"].set(f"{newR:.6f}")







        else:







            k = idx - 1







            self.curR[k] = newR







            self.step_data[k]["R"] = f"{newR:.6f}"







            if self.r_type is None: self.r_type = ["intr"]*len(PRESSURES)







            self.r_type[k] = "man"; self.dragged_flag = True







        self._recompute_from_curves(drag_source="R")







        self._draw_charts()















    def _on_release_rp(self, event):







        if self.dragging and self.dragging[0]=="rp":







            self.dragging = None







            self._mark_dragged()















    def _recompute_from_curves(self, drag_source: str):







        if not self.meta: return







        rho = self.meta["rho"]; m_dry = self.meta["m_dry"]; m_sat = self.meta["m_sat"]







        Vp = self.meta["Vp"]; R0v = self.curR0















        rows = []







        for i,p in enumerate(PRESSURES):







            S = self.curS[i]







            Vv = (Vp*(1-S)) if (S is not None and not math.isnan(S) and Vp>0) else float('nan')







            m = (m_sat - rho*Vv) if not math.isnan(Vv) else float('nan')







            R = self.curR[i]







            Rw = self._try_float(self.vars["R_water"].get()); R_corr = (R * (Rw / R0v)) if (R0v is not None and R0v != 0 and R is not None and not math.isnan(R)) and (Rw is not None) else R







            lam = ((R0v / R_corr) if (R0v is not None and R_corr is not None and not math.isnan(R_corr) and R_corr > 0) else float('nan'))







            msrc = "руч.(drag S)" if drag_source=="S" else ("измерено" if (self.m_type and self.m_type[i]=="meas") else "интерп.")







            rsrc = "руч.(drag R)" if drag_source=="R" else ("измерено" if (self.r_type and self.r_type[i]=="meas") else "интерп.")







            rows.append({"P":p,"m":m,"mSrc":msrc,"R":R_corr,"rSrc":rsrc,







                         "Vv":Vv,"S":S,"S_pct":(S*100.0 if not math.isnan(S) else float('nan')),"lam":lam})















        self.last_rows = rows







        self._render_all(rows, dragged=True)







        self._refresh_inputs_table()















    def _mark_dragged(self): self.dragged_flag = True















    # ---------- Сброс ----------







    def reset_curves(self):







        if self.baseS is None:







            messagebox.showwarning("Нет данных", "Сначала введите исходные данные."); return







        self.curS = self.baseS[:]; self.curR = self.baseR[:]; self.curR0 = self.baseR0







        if self.meta:







            Vp = self.meta["Vp"]; rho = self.meta["rho"]; m_sat = self.meta["m_sat"]







            for i,_ in enumerate(PRESSURES):







                S = self.curS[i]







                Vv = (Vp*(1-S)) if (S is not None and not math.isnan(S) and Vp>0) else float('nan')







                m = (m_sat - rho*Vv) if not math.isnan(Vv) else float('nan')







                self.step_data[i]["m"] = "" if math.isnan(m) else f"{m:.6f}"







                self.step_data[i]["R"] = "" if (self.curR[i] is None or (isinstance(self.curR[i],float) and math.isnan(self.curR[i]))) else f"{self.curR[i]:.6f}"







            self.vars["R0"].set("" if self.curR0 is None else f"{self.curR0:.6f}")







        self.dragged_flag = False







        self._refresh_inputs_table()







        self.calculate(auto=True)















    # ---------- Экспорт ----------







    def export_results(self):







        if not self.last_rows or not self.meta:







            messagebox.showwarning("Нет данных", "Сначала введите данные и выполните расчёт."); return







        if HAS_XLSX:







            path = filedialog.asksaveasfilename(defaultextension=".xlsx",







                                                filetypes=[("Excel Workbook","*.xlsx")],







                                                title="Сохранить отчёт")







            if not path: return







            try:







                wb = Workbook()







                ws1 = wb.active; ws1.title = "Summary"







                m = self.meta







                ws1.append(["Отчёт капилляриметрии (GUI v4.1, drag)"])







                ws1.append(["Диаметр D, мм", m["D"]])







                ws1.append(["Длина L, мм", m["L"]])







                ws1.append(["Объём образца V_обр, см³", r4(m["Vobr"])])







                ws1.append(["Плотность ρ, г/см³", m["rho"]])







                ws1.append(["Масса сухая m_сух, г", m["m_dry"]])







                ws1.append(["Масса насыщ. m_нас, г", m["m_sat"]])







                ws1.append(["Полный объём пор V_p, см³", r4(m["Vp"])])







                ws1.append(["R₀ (Ом·м) при 100% насыщении", ("" if m["R0"] is None else r4(m["R0"]))])







                ws1.append(["Kво (доли)", ("" if (not hasattr(self,'kvo_val') or math.isnan(self.kvo_val)) else r4(self.kvo_val))])







                ws1.append(["Kво (%)", ("" if (not hasattr(self,'kvo_pct') or math.isnan(self.kvo_pct)) else r4(self.kvo_pct))])







                ws1.append(["Drag-редактирование", "Да" if self.dragged_flag else "Нет"])















                ws2 = wb.create_sheet("Steps")







                ws2.append(["P (атм)","Вкл.","Масса (г)","R (Ом)"])







                for d in self.step_data:







                    ws2.append([d["P"], 1 if d["en"] else 0, d["m"], d["R"]])















                ws3 = wb.create_sheet("Results")







                ws3.append(["P (атм)","Масса (г)","Источник m","R (Ом)","Источник R","Vв.выт (см³)","S (доля)","S (%)","λ=R0/R","Pп (доля)","Pн (доля)"])







                for r in self.last_rows:







                    def fmt(v):







                        if v is None: return ""







                        if isinstance(v,float) and (math.isnan(v) or math.isinf(v)): return ""







                        return round(v,6) if isinstance(v,float) else v







                    ws3.append([r["P"], fmt(r["m"]), r["mSrc"], fmt(r["R"]), r["rSrc"],
                                fmt(r["Vv"]), fmt(r["S"]), fmt(r["S_pct"]), fmt(r["lam"]),
                                fmt(r.get("Pп")), fmt(r.get("Pн"))])




















                wb.save(path); messagebox.showinfo("Готово", f"Отчёт сохранён:\n{path}")







            except Exception as e:







                messagebox.showerror("Ошибка", f"Не удалось сохранить Excel:\n{e}")







        else:







            path = filedialog.asksaveasfilename(defaultextension=".csv",







                                                filetypes=[("CSV","*.csv")],







                                                title="Сохранить отчёт (CSV)")







            if not path: return







            try:







                import csv







                with open(path, "w", newline="", encoding="utf-8") as f:







                    w = csv.writer(f, delimiter=";")







                    m = self.meta







                    w.writerow(["Summary"])







                    w.writerow(["Диаметр D, мм", m["D"]]); w.writerow(["Длина L, мм", m["L"]])







                    w.writerow(["Объём образца V_обр, см³", r4(m["Vobr"])])







                    w.writerow(["Плотность ρ, г/см³", m["rho"]])







                    w.writerow(["Масса сухая m_сух, г", m["m_dry"]]); w.writerow(["Масса насыщ. m_нас, г", m["m_sat"]])







                    w.writerow(["Полный объём пор V_p, см³", r4(m["Vp"])])







                    w.writerow(["R₀ (Ом·м) при 100% насыщении", ("" if m["R0"] is None else r4(m["R0"]))])







                    w.writerow(["Kво (доли)", "" if (not hasattr(self,'kvo_val') or math.isnan(self.kvo_val)) else r4(self.kvo_val)])







                    w.writerow(["Kво (%)", "" if (not hasattr(self,'kvo_pct') or math.isnan(self.kvo_pct)) else r4(self.kvo_pct)])







                    w.writerow(["Drag-редактирование", "Да" if self.dragged_flag else "Нет"])







                    w.writerow([]); w.writerow(["Steps"])







                    w.writerow(["P (атм)","Вкл.","Масса (г)","R (Ом)"])







                    for d in self.step_data:







                        w.writerow([d["P"], 1 if d["en"] else 0, d["m"], d["R"]])







                    w.writerow([]); w.writerow(["Results"])







                    w.writerow(["P (атм)","Масса (г)","Источник m","R (Ом)","Источник R","Vв.выт (см³)","S (доля)","S (%)","λ=R0/R"])







                    for r in self.last_rows:







                        def fmt(v):







                            if v is None: return ""







                            if isinstance(v,float) and (math.isnan(v) or math.isinf(v)): return ""







                            return f"{v:.6f}" if isinstance(v,float) else v







                        w.writerow([r["P"], fmt(r["m"]), r["mSrc"], fmt(r["R"]), r["rSrc"],







                                    fmt(r["Vv"]), fmt(r["S"]), fmt(r["S_pct"]), fmt(r["lam"])])







                messagebox.showinfo("Готово", f"CSV сохранён:\n{path}")







            except Exception as e:







                messagebox.showerror("Ошибка", f"Не удалось сохранить CSV:\n{e}")















# ---------- Запуск ----------







if __name__ == "__main__":







    app = App()







    app.mainloop()


# ===============================
# БЛОК РАСЧЁТА Пористости Pп и Водонасыщенности Pн по Арчи (a = 1)
# ===============================
# Формулы Арчи (a=1):
#   F = a / φ^m  = 1 / φ^m           — фактор формации
#   I = S_w^{-n}                      — индекс насыщения
#   φ = (a * R_w / R_0)^{1/m}        — пористость (при известном R0 и Rw)
#   S_w = ((a * R_w) / (φ^m * R_t))^{1/n}  — водонасыщенность (при известном Rt, Rw, φ)
#
# Обозначения:
#   φ  — пористость (доли), здесь: Pп
#   S_w — водонасыщенность (доли), здесь: Pн
#   m — показатель цементации
#   n — показатель насыщения
#   Rw — удельное сопротивление пластовой/фильтрационной воды (Ом·м или Ом, консистентно с R)
#   R0 — сопротивление образца при 100% водонасыщении (Ом)
#   Rt — истинное сопротивление образца/пласта при текущем насыщении (Ом)
#
# Функции ниже аккуратно обрабатывают как скаляры, так и numpy-массивы.
try:
    import numpy as _np  # не требуем numpy жёстко — блок работает и со скалярами
except Exception:  # pragma: no cover
    _np = None

def _to_array(x):
    """Внутренняя утилита: приводим к массиву, если доступен numpy; иначе оставляем как есть."""
    if _np is None:
        return x
    return _np.asarray(x, dtype=float)

# === Формулы Арчи (a = 1) ===
# φ = (a * Rw / R₀)^(1/m) — пористость (Pп)
# S_w = ((a * Rw) / (φ^m * Rt))^(1/n) — водонасыщенность (Pн)
# I = Rt / R₀ — индекс насыщения (безразмерный)
def archie_porosity_Pp(R0, Rw, m, a=1.0):
    """Пористость Pп (φ) по Арчи при заданных R0, Rw и m (a=1 по умолч.).
    Возвращает значения в долях (0..1).

    φ = (a * Rw / R0) ** (1/m)
    """
    if R0 is None or Rw is None or m is None:
        return _np.nan if _np is not None else float('nan')
    xR0 = _to_array(R0); xRw = _to_array(Rw); xm = _to_array(m); xa = _to_array(a)
    # защита от деления на ноль/отрицательных входов
    if _np is not None:
        with _np.errstate(divide='ignore', invalid='ignore'):
            phi = (xa * xRw / xR0) ** (1.0 / xm)
            return phi
    else:
        if R0 == 0:
            return float('nan')
        return (a * Rw / R0) ** (1.0 / m)

def archie_Sw_Pn(Rt, Rw, phi, m, n, a=1.0):
    """Водонасыщенность Pн (S_w) по Арчи при заданных Rt, Rw, φ, m, n (a=1 по умолч.).
    Возвращает значения в долях (0..1).

    S_w = ((a * Rw) / (φ**m * Rt)) ** (1/n)
    """
    if Rt is None or Rw is None or phi is None or m is None or n is None:
        return _np.nan if _np is not None else float('nan')
    xRt = _to_array(Rt); xRw = _to_array(Rw); xphi = _to_array(phi); xm = _to_array(m); xn = _to_array(n); xa = _to_array(a)
    if _np is not None:
        with _np.errstate(divide='ignore', invalid='ignore'):
            denom = (xphi ** xm) * xRt
            Sw = (xa * xRw / denom) ** (1.0 / xn)
            return Sw
    else:
        denom = (phi ** m) * Rt
        if denom == 0:
            return float('nan')
        return (a * Rw / denom) ** (1.0 / n)

# Удобные алиасы с русскими названиями
calc_Pp_porosity = archie_porosity_Pp
calc_Pn_saturation = archie_Sw_Pn

# Мини-тест (не выполняется при обычном запуске GUI):
# Установите переменную окружения RUN_ARCHIE_DEMO=1 чтобы увидеть пример в консоли.
def _demo_archie():  # pragma: no cover
    try:
        import os
        if os.environ.get("RUN_ARCHIE_DEMO") != "1":
            return
        m_demo, n_demo = 2.0, 2.0
        Rw_demo, R0_demo, Rt_demo = 0.1, 10.0, 50.0
        phi = archie_porosity_Pp(R0_demo, Rw_demo, m_demo)     # ≈ (0.1/10)^(1/2) = (0.01)^(0.5) = 0.1
        Sw  = archie_Sw_Pn(Rt_demo, Rw_demo, phi, m_demo, n_demo)  # демонстрационный расчёт
        print(f"[ARCHIE DEMO] φ (Pп) = {phi},  S_w (Pн) = {Sw}")
    except Exception as e:
        print("[ARCHIE DEMO] Error:", e)

_demo_archie()
# =============================== Конец блока Арчи ===============================


# ==================== ВСТАВКА: интеграция Pп и Pн в GUI (и расчёт m из данных) ====================
def _arch_safe_float(x):
    try:
        import math
        v = float(x)
        return v if math.isfinite(v) else float('nan')
    except Exception:
        return float('nan')

try:
    _ORIG__render_results = App._render_results
except Exception:
    _ORIG__render_results = None

def _render_results_with_archie(self, rows):
    # Сначала запускаем оригинальную отрисовку
    if _ORIG__render_results is not None:
        _ORIG__render_results(self, rows)

    # --- Расчёт m из данных (a=1), затем Pп и Pн по Арчи ---
    import math
    import numpy as np

    if not rows or not isinstance(rows, list):
        return

    # Достаём Rw, R0, φ (Kpor/Kpor_eff)
    Rw = _arch_safe_float(self.vars.get("R_water").get() if "R_water" in self.vars else None)
    # Предпочтительно берём curR0, иначе ищем точку S≈1
    R0 = _arch_safe_float(getattr(self, "curR0", float('nan')))
    if not (R0 and math.isfinite(R0)) or R0 <= 0:
        try:
            Sw_arr = np.array([_arch_safe_float(r.get("S", r.get("S_pct", float('nan')))/100 if "S_pct" in r else r.get("S", float('nan'))) for r in rows], dtype=float)
            R_arr  = np.array([_arch_safe_float(r.get("R", float('nan'))) for r in rows], dtype=float)
            idx = np.nanargmin(np.abs(Sw_arr - 1.0))
            R0 = float(R_arr[idx])
        except Exception:
            R0 = float('nan')

    # Пористость — берём эффективную, если есть; иначе обычную
    phi_input = self.meta.get("Kpor_eff", None)
    if not (isinstance(phi_input, (int,float)) and math.isfinite(phi_input) and phi_input>0):
        phi_input = self.meta.get("Kpor", None)

    phi_input = _arch_safe_float(phi_input)

    # Оценим m: m = ln(a*Rw/R0)/ln(phi). a=1. Если phi_input нет — m посчитать нельзя.
    m_est = float('nan')
    if (Rw and R0 and phi_input) and all(math.isfinite(v) for v in [Rw, R0, phi_input]) and (Rw>0 and R0>0 and 0<phi_input<1):
        try:
            m_est = math.log(Rw / R0) / math.log(phi_input)
        except Exception:
            m_est = float('nan')

    # Если m по phi_input не удалось — попробуем взять m из UI/строк
    if not (m_est and math.isfinite(m_est)):
        m_vals = [ _arch_safe_float(r.get("m")) for r in rows ]
        mv = [v for v in m_vals if v and math.isfinite(v) and v>0]
        if mv:
            m_est = float(np.nanmedian(np.array(mv, dtype=float)))
        else:
            try:
                m_est = _arch_safe_float(self.vars.get("m").get()) if "m" in self.vars else float("nan")
            except Exception:
                m_est = float("nan")

    # Если всё ещё нет m — прекращаем
    if not (m_est and math.isfinite(m_est)):
        return

    # Пористость по Арчи из R0,Rw,m (если нет явного phi_input)
    phi_archie = (Rw / R0) ** (1.0 / m_est)

    # Выберем φ для расчёта: если задана Kpor_eff (phi_input) — берем её, иначе phi_archie
    phi_use = float(phi_input) if (phi_input and math.isfinite(phi_input) and 0 < float(phi_input) < 1) else float(phi_archie)

    # Соберём Rt и n по строкам
    Rt_arr = np.array([_arch_safe_float(r.get("R", float('nan'))) for r in rows], dtype=float)
    n_arr  = np.array([_arch_safe_float(r.get("n", float('nan'))) for r in rows], dtype=float)

    # Если n где-то не проставлен — используем медианное по тем, где есть
    if np.all(~np.isfinite(n_arr)) or np.all(n_arr <= 0):
        # fallback: n=2 если совсем ничего нет
        n_arr = np.full_like(Rt_arr, 2.0, dtype=float)
    else:
        good = np.isfinite(n_arr) & (n_arr>0)
        if np.any(good):
            n_med = float(np.nanmedian(n_arr[good]))
            n_arr[~good] = n_med
        else:
            n_arr = np.full_like(Rt_arr, 2.0, dtype=float)

    # Считаем Pн (Sw) построчно
    with np.errstate(divide='ignore', invalid='ignore'):
        denom = (phi_archie ** m_est) * Rt_arr
        Sw_arch = (Rw / denom) ** (1.0 / n_arr)
    # Ограничим разумно [0..1]
    Sw_arch = np.clip(Sw_arch, 0.0, 1.0)

    # Сохраняем в rows и в last_rows
    for i, r in enumerate(rows):
        try:
            r["Pп"] = float(phi_archie) if math.isfinite(phi_archie) else None
            r["Pн"] = float(Sw_arch[i]) if math.isfinite(Sw_arch[i]) else None
        except Exception:
            pass

    # Обновим self.last_rows, если он есть
    try:
        if hasattr(self, "last_rows") and isinstance(self.last_rows, list) and len(self.last_rows)==len(rows):
            for i in range(len(rows)):
                self.last_rows[i]["Pп"] = rows[i].get("Pп")
                self.last_rows[i]["Pн"] = rows[i].get("Pн")
    except Exception:
        pass

    # Добавим колонки в таблицу Results динамически и проставим значения
    try:
        cols = list(self.tree_res["columns"])
        add_cols = []
        if "Pп" not in cols:
            cols.append("Pп"); add_cols.append(("Pп", "Pп (пористость)", 110))
        if "Pн" not in cols:
            cols.append("Pн"); add_cols.append(("Pн", "Pн (водонасыщ.)", 120))
        if add_cols:
            self.tree_res["columns"] = tuple(cols)
            for key, title, width in add_cols:
                self.tree_res.heading(key, text=title)
                self.tree_res.column(key, width=width, anchor="center")

        # Установим значения в существующие строки
        items = self.tree_res.get_children()
        for item, r in zip(items, rows):
            if "Pп" in self.tree_res["columns"]:
                self.tree_res.set(item, "Pп", f"{r['Pп']:.6f}" if isinstance(r.get("Pп"), float) else "")
            if "Pн" in self.tree_res["columns"]:
                self.tree_res.set(item, "Pн", f"{r['Pн']:.6f}" if isinstance(r.get("Pн"), float) else "")
    except Exception as e:
        print("Арчи(Pп/Пн) GUI update error:", e)

    # Сохраним служебные значения на будущее (например, для экспорта)
    self.archie_phi = float(phi_archie) if math.isfinite(phi_archie) else None
    self.archie_m   = float(m_est) if math.isfinite(m_est) else None
    self.archie_R0  = float(R0) if (R0 and math.isfinite(R0)) else None
    self.archie_Rw  = float(Rw) if (Rw and math.isfinite(Rw)) else None

# Переназначаем метод
if _ORIG__render_results is not None:
    App._render_results = _render_results_with_archie
# ==================== КОНЕЦ ВСТАВКИ ====================

